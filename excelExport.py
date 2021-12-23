import openpyxl

# creating the DataFrame
class excelExport:
    def __init__(self, data,):
        self.data = data

    def evaluate(self):
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "Continua"
        hoja.append(('Mochila Binaria', 'Ascenso de la colina','','','','','Algoritmo modificado','','','',''))#'Ascenso a la colina '))
        hoja.append(('', 'Media','Desviación','Peor','Mejor','Tiempo','Media','Desviación','Peor','Mejor','Tiempo'))#,'Media','Desviación','Mejor','Peor','Tiempo'))
        hoja.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
        hoja.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
        hoja.merge_cells(start_row=1, start_column=12, end_row=1, end_column=16)
        hoja.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        auxProblem = self.data[0][0]
        row = []
        row.append(auxProblem)
        for solution in self.data:
            problem = solution[0]
            if auxProblem == problem:
                for i in range(1,len(solution)):
                    row.append(solution[i])
            else:
                hoja.append(row)
                row = []
                row.append(problem)
                auxProblem = problem
                for i in range(1,len(solution)):
                    row.append(solution[i])
        hoja.append(row)
        wb.save('mochila.xlsx')
